#!/usr/bin/env python3
"""
채권 시장 모니터 — HTML 대시보드
평일 하루 2회(오전 9시·오후 3시) 자동 업데이트
"""

import html as _html, json, os, re, subprocess, sys
from datetime import date, datetime, timedelta, timezone

try:
    import requests
except ImportError:
    sys.exit("pip3 install requests 를 먼저 실행하세요.")

# ── 설정 ────────────────────────────────────────────────────────────
OUTPUT     = os.path.join(os.path.dirname(os.path.abspath(__file__)), "index.html")
CACHE_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "kofia_cache.json")
def _ytd_biz_days():
    today = date.today()
    jan1  = date(today.year, 1, 1)
    d, cnt = jan1, 0
    while d <= today:
        if d.weekday() < 5:
            cnt += 1
        d += timedelta(days=1)
    return cnt + 5  # 여유분

LOOKBACK_DAYS = max(30, _ytd_biz_days())

KOFIA_URL = "https://www.kofiabond.or.kr/proframeWeb/XMLSERVICES/"
KOFIA_HDR = {
    "User-Agent":   "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) Chrome/124.0",
    "Content-Type": "text/xml; charset=utf-8",
    "Referer":      "https://www.kofiabond.or.kr/",
}

# 한전 포함 전체 공기업 목록
KEPCO = "한국전력공사"
OTHER_ORGS = [
    "한국도로공사","한국가스공사","한국수자원공사","한국토지주택공사",
    "인천국제공항공사","한국주택금융공사","한국장학재단","한국농어촌공사",
    "한국산업단지공단","한국자산관리공사","국가철도공단","부산항만공사",
]
ALL_ORGS = [KEPCO] + OTHER_ORGS

# KOFIA val키 → 개월수
VAL_MONTHS = {
    "val1":3,"val2":6,"val3":9,"val4":12,"val5":18,
    "val6":24,"val7":30,"val8":36,"val9":48,"val10":60,
    "val11":84,"val12":120,
}

# 민평금리 3사 (나이스피앤아이·한국자산평가·KIS자산평가)
THREE_COMPANIES = {"나이스피앤아이", "한국자산평가", "KIS자산평가"}

def avg3(vals: list):
    v = [x for x in vals if x is not None]
    return round(sum(v) / len(v), 3) if len(v) >= 2 else (v[0] if v else None)

# ── 날짜 유틸 ────────────────────────────────────────────────────────

def last_biz(d=None):
    d = d or date.today()
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d

def prev_biz(d=None):
    d = (d or date.today()) - timedelta(days=1)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d

def biz_days_list(n=20):
    days, d = [], last_biz()
    while len(days) < n:
        if d.weekday() < 5:
            days.append(d)
        d -= timedelta(days=1)
    return list(reversed(days))

def issue_tenor_label(iss_dt: str, mat_dt: str) -> str:
    """발행일~만기일로 실제 발행 만기 계산 → '3년', '5년' 등"""
    if len(iss_dt) != 8 or len(mat_dt) != 8:
        return ""
    try:
        iss = date(int(iss_dt[:4]), int(iss_dt[4:6]), int(iss_dt[6:]))
        mat = date(int(mat_dt[:4]), int(mat_dt[4:6]), int(mat_dt[6:]))
    except ValueError:
        return ""
    total_m = (mat.year - iss.year) * 12 + (mat.month - iss.month)
    if mat.day < iss.day:
        total_m -= 1
    std = [3, 6, 12, 18, 24, 36, 48, 60, 84, 120]
    nearest = min(std, key=lambda x: abs(x - total_m))
    if abs(nearest - total_m) > 3:
        return f"{total_m // 12}년{total_m % 12}개월" if total_m >= 12 else f"{total_m}개월"
    if nearest < 12:
        return f"{nearest}개월"
    if nearest % 12 == 0:
        return f"{nearest // 12}년"
    return f"{nearest // 12}년{nearest % 12}개월"

def closest_val_key(remain_code: str):
    """잔존기간 코드(YYMMDD) → 가장 가까운 val키. 10년 초과 None."""
    if not remain_code or len(remain_code) < 4:
        return None
    try:
        yy, mm = int(remain_code[:2]), int(remain_code[2:4])
        total = yy * 12 + mm
    except ValueError:
        return None
    if total < 1 or total > 150:
        return None
    return min(VAL_MONTHS, key=lambda k: abs(VAL_MONTHS[k] - total))

# ── ProFrame 공통 ────────────────────────────────────────────────────

def kofia_post(svc, fn, fields: dict) -> str:
    dto_name = list(fields.keys())[0]
    inner    = "".join(f"  <{k}>{v}</{k}>\n" for k, v in fields[dto_name].items())
    xml = (f'<?xml version="1.0" encoding="utf-8"?>\n<message>\n'
           f'  <proframeHeader>\n'
           f'    <pfmAppName>BIS-KOFIABOND</pfmAppName>\n'
           f'    <pfmSvcName>{svc}</pfmSvcName>\n'
           f'    <pfmFnName>{fn}</pfmFnName>\n'
           f'  </proframeHeader>\n  <systemHeader></systemHeader>\n'
           f'  <{dto_name}>\n{inner}  </{dto_name}>\n</message>')
    try:
        r = requests.post(KOFIA_URL, data=xml.encode("utf-8"), headers=KOFIA_HDR, timeout=15)
        return r.content.decode("utf-8", errors="ignore")
    except Exception as e:
        print(f"    KOFIA 오류: {e}")
        return ""

def kofia_items(txt, dto="BISComDspDatDTO"):
    return re.findall(f"<{dto}>(.*?)</{dto}>", txt, re.DOTALL)

def pval(row, k):
    m = re.search(f"<{k}>(.*?)</{k}>", row)
    return m.group(1).strip() if m else ""

# ── 1. 국고채·한전채 민평금리 이력 ──────────────────────────────────

def fetch_ktb_day(yyyymmdd):
    """국고채 3사 평균 3년물 → {"국고채": float}"""
    txt = kofia_post("BISBndSrtPrcSrchSO", "selectDay",
                     {"BISBndSrtPrcDayDTO": {
                         "standardDt": yyyymmdd, "reportCompCd": "A10000",
                         "applyGbCd": "C02",
                         "val1":"Y","val2":"Y","val3":"Y","val4":"Y","val5":"Y",
                     }})
    company_vals = []
    for row in kofia_items(txt, "BISBndSrtPrcDayDTO"):
        if (pval(row,"largeCategoryMrk")=="국채" and
                pval(row,"typeNmMrk")=="국고채권" and
                pval(row,"koreanShotNm") in THREE_COMPANIES):
            v = pval(row, "val8")
            if v:
                try: company_vals.append(float(v))
                except ValueError: pass
    result = avg3(company_vals)
    return {"국고채": result} if result else {}

def fetch_org_minp(yyyymmdd):
    """공기업별 3사 평균 전 만기 → {org: {val_key: rate}}"""
    txt = kofia_post("BISBndSrtPrcSrchSO", "selectDay2",
                     {"BISBndSrtPrcDayDTO": {
                         "standardDt": yyyymmdd, "reportCompCd": "A10000",
                         "applyGbCd": "C02",
                         "val1":"Y","val2":"Y","val3":"Y","val4":"Y","val5":"Y",
                     }})
    # 3사별 값 수집
    raw = {}  # {org: {vk: [val, ...]}}
    for row in kofia_items(txt, "BISBndSrtPrcDayDTO"):
        org  = pval(row, "typeNmMrk")
        comp = pval(row, "koreanShotNm")
        if comp not in THREE_COMPANIES:
            continue
        for vk in VAL_MONTHS:
            v = pval(row, vk)
            if v:
                try:
                    raw.setdefault(org, {}).setdefault(vk, []).append(float(v))
                except ValueError:
                    pass
    # 3사 평균 계산
    result = {}
    for org, vk_map in raw.items():
        avgs = {vk: avg3(vals) for vk, vals in vk_map.items() if avg3(vals) is not None}
        if avgs:
            result[org] = avgs
    return result

def collect_kofia_history():
    print("[ KOFIA 국고채·한전채 민평금리 수집 ]")
    cache = {}
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, encoding="utf-8") as f:
                cache = json.load(f)
        except Exception:
            cache = {}

    biz_list  = biz_days_list(LOOKBACK_DAYS)
    need = [d for d in biz_list
            if not (cache.get(d.strftime("%Y-%m-%d"), {}).get("국고채")
                    and cache.get(d.strftime("%Y-%m-%d"), {}).get("한전채"))]

    fetched = 0
    for d in sorted(need, reverse=True):
        if fetched >= 30:          # 한 번에 최대 30일 백필
            break
        dk = d.strftime("%Y-%m-%d")
        ds = d.strftime("%Y%m%d")
        data = {}
        data.update(fetch_ktb_day(ds))
        org_data = fetch_org_minp(ds)
        kepco = org_data.get(KEPCO, {})
        if kepco.get("val8"):
            data["한전채"] = kepco["val8"]
        if data.get("국고채") and data.get("한전채"):
            cache[dk] = data
            print(f"  ✓ {dk}: 국고={data['국고채']}%, 한전={data['한전채']}%")
            fetched += 1

    # 캐시 보관: 당해 연도 1월 1일 이후
    cutoff = date(date.today().year, 1, 1).strftime("%Y-%m-%d")
    cache  = {k: v for k, v in cache.items() if k >= cutoff and v.get("국고채") and v.get("한전채")}
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)

    chart  = {d.strftime("%Y-%m-%d"): cache[d.strftime("%Y-%m-%d")]
              for d in biz_list if cache.get(d.strftime("%Y-%m-%d"))}
    latest = {}
    for d in reversed(biz_list):
        dk = d.strftime("%Y-%m-%d")
        if cache.get(dk):
            latest = {**cache[dk], "날짜": dk}
            break
    return chart, latest

# ── 2. 발행실적 수집 (한전 포함 전체) ───────────────────────────────

def collect_kepco_rates_ytd():
    """KOFIA에서 올해 전력채 발행금리 수집 (차트용, 민평 조회 없음)"""
    print("[ KOFIA 전력채 연초~현재 발행금리 수집 ]")
    today  = last_biz()
    jan1   = date(today.year, 1, 1)
    start  = jan1.strftime("%Y%m%d")
    end    = today.strftime("%Y%m%d")

    txt = kofia_post("BISIssInfoSntcSrchSO", "list",
                     {"BISComDspDatDTO": {
                         "val1": "ISS", "val2": start, "val3": end,
                         "val4": "3", "val5": "", "val6": "", "val7": "",
                     }})
    rows = kofia_items(txt, "BISComDspDatDTO")
    acc  = {}   # {date: [total_원, weighted_sum]}
    for row in rows:
        name = pval(row, "val1")
        if KEPCO not in name: continue
        if "MBS" in name or "유동화" in name: continue
        iss_dt = pval(row, "val3")
        amount = pval(row, "val6")
        coupon = pval(row, "val9")
        iss_fmt = f"{iss_dt[:4]}-{iss_dt[4:6]}-{iss_dt[6:]}" if len(iss_dt)==8 else iss_dt
        try:
            amt  = int(amount) * 100_000_000
            rate = float(coupon)
        except (ValueError, TypeError):
            continue
        if iss_fmt not in acc:
            acc[iss_fmt] = [0, 0.0]
        acc[iss_fmt][0] += amt
        acc[iss_fmt][1] += amt * rate

    rate_by_date = {d: round(v[1]/v[0], 3) for d, v in acc.items() if v[0]}
    amt_by_date  = {d: v[0] // 100_000_000   for d, v in acc.items() if v[0]}
    print(f"  → 전력채 발행금리 {len(rate_by_date)}일치 수집")
    return rate_by_date, amt_by_date


def _mofe_fetch(url, params=None):
    """curl로 mofe.go.kr 요청 (Python requests는 IP 차단됨)"""
    full_url = url
    if params:
        qs = "&".join(f"{k}={v}" for k, v in params.items())
        full_url = f"{url}?{qs}"
    try:
        res = subprocess.run(
            ["curl", "-s", "-L", "--max-time", "20", "-k",
             "-H", "User-Agent: Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
             "-H", "Accept: text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
             "-H", "Accept-Language: ko-KR,ko;q=0.9,en-US;q=0.8",
             "-H", "Referer: https://mofe.go.kr/",
             full_url],
            capture_output=True, text=True, timeout=25
        )
        return res.stdout
    except Exception as e:
        print(f"    curl 오류: {e}")
        return ""


def collect_ktb3y_rates_ytd():
    """기재부 mofe.go.kr에서 올해 국고채 3년물 경쟁입찰 낙찰금리 수집"""
    print("[ 기재부 국고채 3년물 경쟁입찰 낙찰금리 수집 ]")
    LIST_URL   = "https://mofe.go.kr/st/fnancstats/ktb50201.do"
    DETAIL_URL = "https://mofe.go.kr/st/fnancstats/updateTbFnancstatsView.do"
    year = str(date.today().year)
    serials_3y = []

    try:
        for page_idx in range(1, 8):
            html_text = _mofe_fetch(LIST_URL, {"pageIndex": page_idx})
            if not html_text or len(html_text) < 500:
                print(f"  페이지 {page_idx} 응답 없음")
                break

            rows_html = re.findall(r'<tr[^>]*>(.*?)</tr>', html_text, re.DOTALL)
            any_this_year = False

            for row in rows_html:
                sid_m = re.search(r'(MOSF_\d+)', row)
                if not sid_m:
                    continue
                sid = sid_m.group(1)

                t_m = re.search(r'<td[^>]*class="[^"]*(?:tit|title|subject)[^"]*"[^>]*>(.*?)</td>', row, re.DOTALL)
                title = re.sub(r'<[^>]+>', '', t_m.group(1)).strip() if t_m else ""

                d_m = re.search(r'(\d{4}\.\d{2}\.\d{2})', row)
                reg_date = d_m.group(1) if d_m else ""

                if reg_date.startswith(year):
                    any_this_year = True
                if not reg_date.startswith(year):
                    continue

                if "3년물" in title and "경쟁입찰" in title and "결과" in title:
                    if sid not in serials_3y:
                        serials_3y.append(sid)
                        print(f"  발견: {reg_date} | {title[:45]}")

            if not any_this_year and page_idx > 1:
                break

    except Exception as e:
        print(f"  목록 조회 오류: {e}")
        return {}, {}

    if not serials_3y:
        print("  3년물 경쟁입찰 결과 없음")
        return {}, {}

    rate_by_date = {}
    amt_by_date  = {}

    for sid in serials_3y:
        try:
            content = _mofe_fetch(DETAIL_URL, {"searchSn": sid, "type": "ktb50201"})
            if not content:
                continue

            cn_m = (re.search(r'<input[^>]*id="cn"[^>]*value="([^"]*)"', content) or
                    re.search(r'<input[^>]*value="([^"]*)"[^>]*id="cn"', content))
            if not cn_m:
                continue
            cn = _html.unescape(cn_m.group(1)).replace('\xa0', ' ')

            iss_m = re.search(r"발행일\s*[:：]\s*'?(\d{2})\.(\d{1,2})\.(\d{1,2})", cn)
            if not iss_m:
                continue
            iss_date = (f"20{iss_m.group(1)}-"
                        f"{int(iss_m.group(2)):02d}-"
                        f"{int(iss_m.group(3)):02d}")

            rm = re.search(r"가중평균낙찰금리\s*[:：]\s*([0-9]+\.[0-9]+)", cn)
            if not rm:
                continue
            rate = float(rm.group(1))

            am  = re.search(r"낙찰금액\s*[:：]\s*([0-9,]+)\s*억", cn)
            amt = int(am.group(1).replace(",", "")) if am else 0

            rate_by_date[iss_date] = rate
            if amt:
                amt_by_date[iss_date] = amt
            print(f"  {iss_date}: {rate}%  {amt}억원")

        except Exception as e:
            print(f"  {sid} 상세조회 오류: {e}")

    print(f"  → 국고채 3년 발행금리 {len(rate_by_date)}일치 수집")
    return rate_by_date, amt_by_date


def collect_issuance(days=7):
    print(f"[ KOFIA 발행실적 수집 (최근 {days}일, 한전 포함) ]")
    end_dt   = last_biz()
    start_dt = end_dt - timedelta(days=days + 5)

    txt = kofia_post("BISIssInfoSntcSrchSO", "list",
                     {"BISComDspDatDTO": {
                         "val1": "ISS",
                         "val2": start_dt.strftime("%Y%m%d"),
                         "val3": end_dt.strftime("%Y%m%d"),
                         "val4": "3",
                         "val5": "", "val6": "", "val7": "",
                     }})
    rows = kofia_items(txt, "BISComDspDatDTO")
    cnt  = re.search(r"<dbio_total_count_>(\d+)", txt)
    print(f"  특수채 전체 {cnt.group(1) if cnt else '?'}건 처리 중")

    result = []
    for row in rows:
        name   = pval(row, "val1")
        iss_dt = pval(row, "val3")
        mat_dt = pval(row, "val4")
        remain = pval(row, "val5")
        amount = pval(row, "val6")
        coupon = pval(row, "val9")

        # 한전 또는 기타 공기업 판별
        if KEPCO in name:
            org = KEPCO
        else:
            org = next((o for o in OTHER_ORGS if o in name), None)
        if not org:
            continue
        if "MBS" in name or "유동화" in name:
            continue

        tenor   = issue_tenor_label(iss_dt, mat_dt)
        iss_fmt = f"{iss_dt[:4]}-{iss_dt[4:6]}-{iss_dt[6:]}" if len(iss_dt)==8 else iss_dt
        mat_fmt = f"{mat_dt[:4]}-{mat_dt[4:6]}-{mat_dt[6:]}" if len(mat_dt)==8 else ""
        result.append({
            "org":      org,
            "is_kepco": org == KEPCO,
            "name":     name[:35],
            "date":     iss_fmt,
            "mat_dt":   mat_fmt,
            "tenor":    tenor,
            "remain":   remain,
            "amount":   f"{int(amount):,}" if amount.isdigit() else amount,
            "rate":     coupon,
            "minp":     None,
            "minp_date": "",
        })

    result.sort(key=lambda x: (-int(x["date"].replace("-", "")), x["org"]))

    # 민평금리 매핑 (발행일 전 영업일 기준)
    dates_needed = {}
    for r in result:
        try:
            iss_d = date.fromisoformat(r["date"])
        except ValueError:
            continue
        ld = prev_biz(iss_d)
        dates_needed.setdefault(ld.strftime("%Y%m%d"), ld.strftime("%m/%d"))

    minp_by_date = {}
    for ds, lbl in dates_needed.items():
        print(f"  민평금리 조회: {ds[:4]}-{ds[4:6]}-{ds[6:]}")
        minp_by_date[ds] = fetch_org_minp(ds)

    for r in result:
        try:
            iss_d = date.fromisoformat(r["date"])
        except ValueError:
            continue
        ld  = prev_biz(iss_d)
        ds  = ld.strftime("%Y%m%d")
        lbl = ld.strftime("%m/%d")
        org_minp = minp_by_date.get(ds, {}).get(r["org"], {})
        vk = closest_val_key(r["remain"])
        if vk and vk in org_minp:
            r["minp"]      = org_minp[vk]
            r["minp_date"] = lbl

    print(f"  → 발행실적 총 {len(result)}건 (한전 {sum(1 for r in result if r['is_kepco'])}건 포함)")
    return result

# ── 3. 차입금 현황 관리 ─────────────────────────────────────────────

DEBT_FILE    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "debt_positions.json")
DEBT_HISTORY = os.path.join(os.path.dirname(os.path.abspath(__file__)), "debt_history.json")
EXCEL_FILE   = os.path.join(os.path.dirname(os.path.abspath(__file__)), "업로드 파일(26.5.19).xlsx")
DEBT_BOND_CATS = {"전력채", "단기사채", "외화채권"}

def debt_classify(name):
    if name in {"전력채","단기사채","외화채권","중장기기업어음","은행차입"}:
        return name
    if name in {"나이지리아","남북협력기금","농어촌융자금"}:
        return "기타"
    return None

def init_debt_from_excel():
    try:
        import openpyxl
    except ImportError:
        print("  openpyxl 없음 — pip3 install openpyxl"); return []
    if not os.path.exists(EXCEL_FILE):
        print(f"  엑셀 파일 없음: {EXCEL_FILE}"); return []
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['raw']
    positions = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True), start=3):
        name = row[4]; bal = row[9]; cat = debt_classify(name)
        if not cat or not bal or bal <= 0: continue
        mat = row[16]; iss = row[15]; rate = row[8]
        def _d(v):
            if hasattr(v,"date"): return v.date().isoformat()
            return v.isoformat() if v else None
        mat_s, iss_s = _d(mat), _d(iss)
        uid = f"{cat}|{iss_s}|{mat_s}|{int(bal)}|r{row_idx}"
        positions.append({"id":uid,"category":cat,"amount":int(bal),
            "issuance_date":iss_s,"maturity_date":mat_s,
            "rate":float(rate) if isinstance(rate,(int,float)) and rate else None,
            "source":"excel"})
    print(f"  엑셀 초기화: {len(positions)}건 로드")
    return positions

def load_debt_positions():
    if os.path.exists(DEBT_FILE):
        with open(DEBT_FILE, encoding="utf-8") as f: return json.load(f)
    print("[ 차입금 초기화: 엑셀에서 로드 ]")
    positions = init_debt_from_excel()
    save_debt_positions(positions)
    return positions

def save_debt_positions(positions):
    with open(DEBT_FILE,"w",encoding="utf-8") as f:
        json.dump(positions, f, ensure_ascii=False, indent=2)

def get_debt_summary(positions):
    from collections import defaultdict
    by = defaultdict(int)
    for p in positions: by[p["category"]] += p["amount"]
    사채 = sum(by[c] for c in DEBT_BOND_CATS)
    사채외 = sum(v for c,v in by.items() if c not in DEBT_BOND_CATS)
    return {"전력채":by["전력채"],"단기사채":by["단기사채"],"외화채권":by["외화채권"],
            "중장기기업어음":by["중장기기업어음"],"은행차입":by["은행차입"],"기타":by["기타"],
            "사채":사채,"사채외":사채외,"합계":사채+사채외}

def _load_history():
    if not os.path.exists(DEBT_HISTORY): return {}
    with open(DEBT_HISTORY,encoding="utf-8") as f: return json.load(f)

def save_debt_snapshot(summary, date_str):
    h = _load_history(); h[date_str] = summary
    if len(h) > 90: del h[sorted(h)[0]]
    with open(DEBT_HISTORY,"w",encoding="utf-8") as f: json.dump(h,f,ensure_ascii=False,indent=2)

def get_snapshot(date_str): return _load_history().get(date_str)

SEIBRO_URL = "https://seibro.or.kr/websquare/engine/proworks/callServletService.jsp"
SEIBRO_HDR = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Content-Type": "application/xml; charset=UTF-8",
    "Referer": "https://seibro.or.kr/websquare/control.jsp?w2xPath=/IPORTAL/user/moneyMarke/BIP_CNTS04010V.xml",
    "Origin": "https://seibro.or.kr",
    "X-Requested-With": "XMLHttpRequest",
}
KEPCO_CUSTNO = "1576"

def collect_seibro_stb(start_dt: str, end_dt: str) -> list:
    """SEIBRO에서 한전 단기사채 목록 수집 (날짜 형식: YYYYMMDD)"""
    print(f"  SEIBRO 단기사채 조회: {start_dt[:4]}-{start_dt[4:6]}-{start_dt[6:]} ~ {end_dt[:4]}-{end_dt[4:6]}-{end_dt[6:]}")
    body = (
        f'<reqParam action="issuSecnPListEL1" task="ksd.safe.bip.cnts.MoneyMarke.process.EstpIssuSecnPTask">'
        f'<MENU_NO value="134"/>'
        f'<CMM_BTN_ABBR_NM value="total_search,openall,print,hwp,word,pdf,searchIcon,searchIcon,seach,xls,"/>'
        f'<W2XPATH value="/IPORTAL/user/moneyMarke/BIP_CNTS04010V.xml"/>'
        f'<ISSUCO_CUSTNO value="{KEPCO_CUSTNO}"/>'
        f'<ISIN value=""/>'
        f'<ic_start value="{start_dt}"/>'
        f'<ic_end value="{end_dt}"/>'
        f'<ic_start2 value=""/><ic_end2 value=""/>'
        f'<START_PAGE value="1"/><END_PAGE value="200"/>'
        f'<INDTP_CLSF_NO value=""/>'
        f'</reqParam>'
    )
    try:
        r = requests.post(SEIBRO_URL, data=body.encode("utf-8"), headers=SEIBRO_HDR, timeout=15)
        txt = r.content.decode("utf-8", errors="ignore")
    except Exception as e:
        print(f"    SEIBRO 오류: {e}")
        return []

    items = re.findall(r"<result>(.*?)</result>", txt, re.DOTALL)
    result = []
    for item in items:
        def gv(k):
            m = re.search(f'<{k} value="(.*?)"/>', item)
            return m.group(1).strip() if m else ""
        issu_dt = gv("ISSU_DT")
        xpir_dt = gv("XPIR_DT")
        face_amt = gv("FACE_AMT")
        isin = gv("ISIN_NO") or gv("SCRS_ISIN_CD") or gv("ISIN") or ""
        if not issu_dt or not face_amt:
            continue
        iss_fmt = f"{issu_dt[:4]}-{issu_dt[4:6]}-{issu_dt[6:]}"
        mat_fmt = f"{xpir_dt[:4]}-{xpir_dt[4:6]}-{xpir_dt[6:]}" if len(xpir_dt) == 8 else ""
        try:
            amt = int(face_amt)
        except ValueError:
            continue
        result.append({"issuance_date": iss_fmt, "maturity_date": mat_fmt, "amount": amt, "isin": isin})
    print(f"    → 한전 단기사채 {len(result)}건")
    return result

def update_debt_pm(positions, issuances):
    today_s = last_biz().isoformat()

    # 오늘 발행된 구형(4-pipe) SEIBRO 항목 제거 → ISIN 기반으로 재추가
    old_seibro = [p for p in positions if
                  p.get("source") == "seibro" and
                  p.get("issuance_date") == today_s and
                  p["id"].count("|") == 4]
    if old_seibro:
        print(f"  구형 SEIBRO 항목 제거(재수집): {len(old_seibro)}건")
        positions = [p for p in positions if p not in old_seibro]

    before  = len(positions)
    positions = [p for p in positions if not p["maturity_date"] or p["maturity_date"] > today_s]
    if before-len(positions): print(f"  만기 차감: {before-len(positions)}건")

    existing_ids  = {p["id"] for p in positions}
    existing_keys = {(p["category"], p.get("issuance_date",""), p.get("maturity_date","") or "", p["amount"])
                     for p in positions}
    added = 0

    # KOFIA 전력채 신규 추가
    for iss in issuances:
        if not iss.get("is_kepco"): continue
        try: amt = int(iss["amount"].replace(",","")) * 100_000_000
        except (ValueError,AttributeError): continue
        mat_s = iss.get("mat_dt","") or ""
        key   = ("전력채", iss["date"], mat_s, amt)
        if key in existing_keys: continue
        uid = f"전력채|{iss['date']}|{mat_s}|{amt}|kofia"
        positions.append({"id":uid,"category":"전력채","amount":amt,
            "issuance_date":iss["date"],"maturity_date":mat_s or None,
            "rate":float(iss["rate"]) if iss.get("rate") else None,"source":"kofia"})
        existing_keys.add(key); added += 1
    if added: print(f"  신규 전력채 추가: {added}건")

    # SEIBRO 단기사채 신규 추가 (최근 5영업일 재조회 → 15:00 이후 SEIBRO 등록분 보완)
    today_dt = last_biz()
    lookback = today_dt
    for _ in range(4):          # 5영업일 전까지
        lookback = prev_biz(lookback)
    stb_start = lookback.strftime("%Y%m%d")
    stb_end   = today_dt.strftime("%Y%m%d")
    stb_list  = collect_seibro_stb(stb_start, stb_end)
    stb_added = 0
    for stb in stb_list:
        isin = stb.get("isin", "")
        uid = f"단기사채|{isin}|seibro" if isin else \
              f"단기사채|{stb['issuance_date']}|{stb['maturity_date']}|{stb['amount']}|seibro"
        if uid in existing_ids:
            continue
        if not isin:
            key = ("단기사채", stb["issuance_date"], stb["maturity_date"], stb["amount"])
            if key in existing_keys:
                continue
            existing_keys.add(key)
        positions.append({"id":uid,"category":"단기사채","amount":stb["amount"],
            "issuance_date":stb["issuance_date"],"maturity_date":stb["maturity_date"] or None,
            "rate":None,"source":"seibro"})
        existing_ids.add(uid); stb_added += 1
    if stb_added: print(f"  신규 단기사채 추가: {stb_added}건")

    return positions

def collect_debt(issuances):
    print("[ 차입금 현황 업데이트 ]")
    KST   = timezone(timedelta(hours=9))
    is_pm = datetime.now(KST).hour >= 15
    today = last_biz(); yesterday = prev_biz(today)
    positions = load_debt_positions()
    if is_pm:
        positions = update_debt_pm(positions, issuances)
        save_debt_positions(positions)
        summary = get_debt_summary(positions)
        save_debt_snapshot(summary, today.isoformat())
        as_of  = today.isoformat()
        prev_s = get_snapshot(yesterday.isoformat())
    else:
        as_of  = yesterday.isoformat()
        snap   = get_snapshot(yesterday.isoformat())
        summary = snap if snap else get_debt_summary(positions)
        prev_s  = get_snapshot(prev_biz(yesterday).isoformat())
    return summary, prev_s, as_of, is_pm


# ── 4. 채권 발행 통계 ───────────────────────────────────────────────

def _fetch_issu_stat(start: str, end: str, val4: str):
    """KOFIA에서 기간·종류별 발행 합계(억원) 반환 → (total_억, kepco_억)"""
    txt = kofia_post("BISIssInfoSntcSrchSO", "list",
                     {"BISComDspDatDTO": {
                         "val1": "ISS", "val2": start, "val3": end,
                         "val4": val4, "val5": "", "val6": "", "val7": "",
                     }})
    rows = kofia_items(txt, "BISComDspDatDTO")
    total = kepco = 0
    for row in rows:
        amt = pval(row, "val6")
        if not amt.isdigit(): continue
        total += int(amt)
        if "한국전력" in pval(row, "val1"):
            kepco += int(amt)
    return total, kepco

def collect_issu_stats():
    """올해·작년 채권 발행 통계 수집 — 캐시 포함"""
    print("[ 채권 발행 통계 수집 ]")
    cache = {}
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, encoding="utf-8") as f:
                cache = json.load(f)
        except Exception:
            cache = {}

    today = date.today()
    this_year = str(today.year)
    last_year = str(today.year - 1)
    cache_key  = f"issu_stats_{today.isoformat()}"

    if cache_key in cache:
        print("  캐시 사용")
        return cache[cache_key]

    # 국채(1), 지방채(2), 특수채(3), 금융채(4), 회사채(5) — 단기사채 제외
    CODES = ["1", "2", "3", "4", "5"]
    result = {}
    for year, start, end in [
        (this_year, f"{this_year}0101", today.strftime("%Y%m%d")),
        (last_year, f"{last_year}0101", f"{last_year}1231"),
    ]:
        grand = ktb = speical = kepco = 0
        for code in CODES:
            tot, kep = _fetch_issu_stat(start, end, code)
            grand += tot
            if code == "1": ktb     = tot
            if code == "3": speical = tot; kepco = kep
        result[year] = {
            "전력채": kepco,
            "특수채": speical,
            "국채":   ktb,
            "전체":   grand,
        }
        print(f"  {year}: 전력채 {kepco/10000:.2f}조 / 특수채 {speical/10000:.2f}조 / "
              f"국채 {ktb/10000:.2f}조 / 전체 {grand/10000:.2f}조")

    cache[cache_key] = result
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    return result

def issu_stats_section_html(stats: dict) -> str:
    if not stats: return ""
    today   = date.today()
    this_yr = str(today.year)
    last_yr = str(today.year - 1)

    def fmt조(v): return f"{v/10000:.2f}조"

    def year_data(yr):
        d = stats.get(yr, {})
        kepco   = d.get("전력채", 0)
        special = d.get("특수채", 0)
        ktb     = d.get("국채",   0)
        total   = d.get("전체",   0)
        other_special = special - kepco          # 특수채 중 전력채 제외
        other_total   = total - special - ktb    # 전체 중 특수채·국채 제외
        return kepco, other_special, ktb, other_total, special, total

    def year_html(yr):
        kepco, other_sp, ktb, other_tot, special, total = year_data(yr)
        label = f"{yr}년 ({today.strftime('%m/%d')} 기준)" if yr == this_yr else f"{yr}년 (연간)"

        # ── 상단 강조 카드: 전력채 ──
        sp_pct  = f"{kepco/special*100:.1f}%" if special else "–"
        tot_pct = f"{kepco/total*100:.1f}%"   if total  else "–"
        highlight = f'''
<div style="background:linear-gradient(135deg,#fff7ed,#ffedd5);border:2px solid #f97316;border-radius:14px;padding:20px 24px;margin-bottom:16px;display:flex;align-items:center;gap:24px;flex-wrap:wrap;max-width:544px">
  <div>
    <div style="font-size:.78rem;color:#ea580c;font-weight:600;margin-bottom:4px">⚡ 전력채 연간 발행액</div>
    <div style="font-size:2.2rem;font-weight:700;color:#c2410c">{fmt조(kepco)}</div>
  </div>
  <div style="display:flex;gap:20px;flex-wrap:wrap">
    <div style="text-align:center">
      <div style="font-size:.72rem;color:#92400e">특수채 내 비중</div>
      <div style="font-size:1.3rem;font-weight:700;color:#f97316">{sp_pct}</div>
    </div>
    <div style="text-align:center">
      <div style="font-size:.72rem;color:#92400e">전체 채권 내 비중</div>
      <div style="font-size:1.3rem;font-weight:700;color:#f97316">{tot_pct}</div>
    </div>
  </div>
</div>'''

        # ── 도넛 차트 + 수평 바 ──
        chart_id = f"statChart_{yr}"
        kepco_v   = round(kepco/10000, 2)
        other_v   = round(other_sp/10000, 2)
        ktb_v     = round(ktb/10000, 2)
        othertot_v= round(other_tot/10000, 2)

        def hbar(label, val, total_val, color):
            pct = min(int(val/total_val*100), 100) if total_val else 0
            return f'''
<div style="margin-bottom:10px">
  <div style="display:flex;justify-content:space-between;font-size:.8rem;margin-bottom:3px">
    <span style="color:#475569">{label}</span>
    <span style="font-weight:600;color:{color}">{fmt조(val)} <span style="font-weight:400;color:#94a3b8">({pct}%)</span></span>
  </div>
  <div style="background:#e2e8f0;border-radius:4px;height:10px">
    <div style="background:{color};width:{pct}%;height:10px;border-radius:4px;transition:width .4s"></div>
  </div>
</div>'''

        bars = (hbar("⚡ 전력채",       kepco,     total, "#f97316") +
                hbar("특수채(전력채 외)", other_sp,  total, "#a78bfa") +
                hbar("국채",            ktb,       total, "#3b82f6") +
                hbar("기타(금융채·회사채 등)", other_tot, total, "#94a3b8"))

        total_label = fmt조(total)
        donut_script = f"""
<script>
(function(){{
  var ctx=document.getElementById('{chart_id}').getContext('2d');
  var totalLabel='{total_label}';
  new Chart(ctx,{{
    type:'doughnut',
    data:{{
      labels:['전력채','특수채(전력채외)','국채','기타'],
      datasets:[{{
        data:[{kepco_v},{other_v},{ktb_v},{othertot_v}],
        backgroundColor:['#f97316','#a78bfa','#3b82f6','#cbd5e1'],
        borderWidth:2,borderColor:'#fff',
        hoverOffset:8
      }}]
    }},
    options:{{
      cutout:'62%',
      plugins:{{
        legend:{{position:'bottom',labels:{{font:{{size:11}},padding:10}}}},
        tooltip:{{callbacks:{{label:function(c){{return c.label+': '+c.parsed.toFixed(2)+'조원'}}}}}},
        beforeDraw:undefined
      }}
    }},
    plugins:[{{
      id:'centerText',
      beforeDraw:function(chart){{
        var ca=chart.chartArea,ctx=chart.ctx;
        var cx=(ca.left+ca.right)/2,cy=(ca.top+ca.bottom)/2;
        ctx.restore();
        ctx.font='bold 12px Noto Sans KR,sans-serif';
        ctx.textBaseline='middle';
        ctx.textAlign='center';
        ctx.fillStyle='#64748b';
        ctx.fillText('전체', cx, cy-10);
        ctx.font='bold 14px Noto Sans KR,sans-serif';
        ctx.fillStyle='#0f2a4a';
        ctx.fillText(totalLabel, cx, cy+10);
        ctx.save();
      }}
    }}]
  }});
}})();
</script>"""

        return f'''<div id="stat_{yr}" class="stat-panel">
<p style="font-size:.8rem;color:#64748b;margin-bottom:12px">{label}</p>
{highlight}
<div style="display:grid;grid-template-columns:220px 300px;gap:24px;align-items:center;flex-wrap:wrap">
  <div><canvas id="{chart_id}" style="max-height:220px"></canvas></div>
  <div>{bars}</div>
</div>
{donut_script}
</div>'''

    this_html = year_html(this_yr)
    last_html = year_html(last_yr)

    return f"""
<section>
  <h2 style="border-left-color:#7c3aed">채권 발행 통계</h2>
  <div style="margin-bottom:14px">
    <button onclick="showStat('{this_yr}')" id="btn_{this_yr}"
      style="margin-right:8px;padding:6px 18px;border-radius:20px;border:2px solid #2563eb;background:#2563eb;color:#fff;font-size:.83rem;cursor:pointer;font-weight:600">올해({this_yr})</button>
    <button onclick="showStat('{last_yr}')" id="btn_{last_yr}"
      style="padding:6px 18px;border-radius:20px;border:2px solid #cbd5e1;background:#fff;color:#64748b;font-size:.83rem;cursor:pointer;font-weight:600">작년({last_yr})</button>
  </div>
  {this_html}
  {last_html}
</section>
<script>
function showStat(yr){{
  ['{this_yr}','{last_yr}'].forEach(function(y){{
    document.getElementById('stat_'+y).style.display = y===yr?'block':'none';
    var btn=document.getElementById('btn_'+y);
    if(y===yr){{btn.style.background='#2563eb';btn.style.color='#fff';btn.style.borderColor='#2563eb';}}
    else{{btn.style.background='#fff';btn.style.color='#64748b';btn.style.borderColor='#cbd5e1';}}
  }});
}}
showStat('{this_yr}');
</script>"""

# ── 5. HTML 생성 ────────────────────────────────────────────────────

def build_chart_data(chart, kepco_rate_by_date=None, ktb_rate_by_date=None):
    dates   = sorted(chart.keys())
    labels  = [d[5:].replace("-", "/") for d in dates]
    spreads = [round(chart[d]["한전채"] - chart[d]["국고채"], 4) for d in dates]

    # 금리 y축 범위: 최소 2.0% 고정, 상단은 데이터 기반
    rates = ([chart[d]["국고채"] for d in dates] + [chart[d]["한전채"] for d in dates])
    rates = [r for r in rates if r is not None]
    rate_min = 2.5
    rate_max = round(max(rates) + 0.15, 2) if rates else 4.5

    # 스프레드 y2 범위: 스프레드가 금리선 아래 표시되도록 max를 넓게
    valid    = [s for s in spreads if s is not None]
    raw_max  = max(valid) if valid else 0.40
    sprd_min = 0.0
    sprd_max = round(raw_max * 1.75, 2)  # 스프레드가 차트 하단 ~57% 구간, 금리선 아래 유지

    ds = [
        {"label":"국고채(3년) 민평 (좌)","data":[chart[d]["국고채"] for d in dates],
         "type":"line","borderColor":"#2563eb","backgroundColor":"transparent",
         "borderWidth":2,"pointRadius":2,"pointHoverRadius":6,"tension":0.3,"yAxisID":"y","order":1},
        {"label":"한전채(3년) 민평 (좌)","data":[chart[d]["한전채"] for d in dates],
         "type":"line","borderColor":"#f97316","backgroundColor":"transparent",
         "borderWidth":2,"pointRadius":2,"pointHoverRadius":6,"tension":0.3,"yAxisID":"y","order":2},
        {"label":"스프레드(한전-국고) (우)","data":spreads,
         "type":"line","fill":"start","borderColor":"rgba(20,184,166,0.9)",
         "backgroundColor":"rgba(20,184,166,0.12)","borderWidth":1.5,
         "pointRadius":1.5,"pointHoverRadius":5,"tension":0.3,"yAxisID":"y2","order":3},
        {"label":"전력채 발행금리 (좌)","data":[kepco_rate_by_date.get(d) for d in dates] if kepco_rate_by_date else [None]*len(dates),
         "type":"bar",
         "backgroundColor":"rgba(249,115,22,0.28)","borderColor":"rgba(249,115,22,0.75)",
         "borderWidth":1.5,
         "borderRadius":{"topLeft":4,"topRight":4,"bottomLeft":0,"bottomRight":0},
         "maxBarThickness":18,"yAxisID":"y","order":5},
        {"label":"국고채 3년 발행금리 (좌)","data":[ktb_rate_by_date.get(d) for d in dates] if ktb_rate_by_date else [None]*len(dates),
         "type":"bar",
         "backgroundColor":"rgba(99,102,241,0.28)","borderColor":"rgba(99,102,241,0.75)",
         "borderWidth":1.5,
         "borderRadius":{"topLeft":4,"topRight":4,"bottomLeft":0,"bottomRight":0},
         "maxBarThickness":18,"yAxisID":"y","order":6},
    ]
    return (json.dumps(labels, ensure_ascii=False), json.dumps(ds, ensure_ascii=False),
            sprd_min, sprd_max, rate_min, rate_max)


def issue_row_html(r, show_minp=True):
    """발행실적 1행 HTML"""
    tenor = r["tenor"] or "–"
    # 만기 뱃지 색
    t = tenor
    if any(x in t for x in ["3년","2년6개월"]):
        bc = "badge-3y"
    elif any(x in t for x in ["5년","4년"]):
        bc = "badge-5y"
    else:
        bc = "badge-10y"

    amt_html  = f'{r["amount"]}억원'
    try:
        rate_html = f'{float(r["rate"]):.2f}%'
    except (ValueError, TypeError):
        rate_html = f'{r["rate"]}%' if r["rate"] else "–"

    if r["minp"] is not None:
        minp_html = (f'<span style="color:#0d9488;font-weight:600">{r["minp"]:.3f}%</span>'
                     f'<span style="font-size:.72rem;color:#94a3b8;margin-left:3px">({r["minp_date"]})</span>')
        try:
            diff = round(float(r["rate"]) - r["minp"], 3)
            if diff > 0:
                diff_html = f'<span style="color:#dc2626;font-weight:600">+{diff:.3f}%p</span>'
            elif diff < 0:
                diff_html = f'<span style="color:#2563eb;font-weight:600">{diff:.3f}%p</span>'
            else:
                diff_html = '<span style="color:#94a3b8">0.000%p</span>'
        except (ValueError, TypeError):
            diff_html = "–"
    else:
        minp_html = '<span style="color:#cbd5e1">–</span>'
        diff_html = '<span style="color:#cbd5e1">–</span>'

    return (f'<tr>'
            f'<td style="white-space:nowrap">{r["date"]}</td>'
            f'<td><strong>{r["org"]}</strong></td>'
            f'<td style="font-size:.81rem;color:#64748b;white-space:nowrap">{r["name"]}</td>'
            f'<td style="white-space:nowrap"><span class="badge {bc}">{tenor}</span></td>'
            f'<td style="white-space:nowrap">{amt_html}</td>'
            f'<td class="rate" style="white-space:nowrap">{rate_html}</td>'
            f'<td style="white-space:nowrap">{minp_html}</td>'
            f'<td style="white-space:nowrap">{diff_html}</td>'
            f'</tr>')


def today_section_html(issuances):
    """금일 발행현황 섹션 — 항상 오늘 날짜 표시"""
    today     = last_biz()
    yesterday = prev_biz(today)
    today_str = today.strftime("%Y-%m-%d")
    yest_str  = yesterday.strftime("%Y-%m-%d")
    section_date = today.strftime("%Y년 %m월 %d일")

    today_all = [r for r in issuances if r["date"] == today_str]
    has_today = len(today_all) > 0
    KST = timezone(timedelta(hours=9))
    is_afternoon = datetime.now(KST).hour >= 15

    if has_today:
        display_rows = today_all
        update_note  = ''
    elif is_afternoon:
        display_rows = []
        update_note  = ''
    else:
        display_rows = [r for r in issuances if r["date"] == yest_str]
        update_note  = '⏰ 금일 발행 데이터는 오후 3시 업데이트 예정 · 현재 전일({}) 기준'.format(yesterday.strftime("%m/%d"))

    kepco_rows  = [r for r in display_rows if r["is_kepco"]]
    others_rows = [r for r in display_rows if not r["is_kepco"]]

    def mini_table(rows, empty_msg):
        if not rows:
            return f'<p style="color:#94a3b8;font-size:.87rem;padding:14px 0">{empty_msg}</p>'
        thead = ('<thead><tr>'
                 '<th>발행일</th><th>발행기관</th><th>종목명</th>'
                 '<th>만기</th><th>발행금액</th><th>발행금리</th>'
                 '<th>전일 민평</th><th>발행-민평</th>'
                 '</tr></thead>')
        tbody = "\n".join(issue_row_html(r) for r in rows)
        return f'<div class="table-wrap"><table>{thead}<tbody>{tbody}</tbody></table></div>'

    kepco_html  = mini_table(kepco_rows,  "발행 없음")
    others_html = mini_table(others_rows, "발행 없음")

    note_html = f'<p style="font-size:.82rem;color:#64748b;margin-bottom:14px">{update_note}</p>' if update_note else ''
    return f"""
<section>
  <h2 style="border-left-color:#f97316">채권 발행현황</h2>
  {note_html}
  <div class="today-grid">
    <div>
      <h3 class="sub-h3" style="color:#f97316">⚡ 한국전력공사</h3>
      {kepco_html}
    </div>
    <div>
      <h3 class="sub-h3" style="color:#2563eb">🏛 타기관 공사채</h3>
      {others_html}
    </div>
  </div>
</section>"""


def recent_section_html(issuances):
    """최근 7일 전체 발행실적"""
    if not issuances:
        return '<tr><td colspan="8" style="text-align:center;color:#94a3b8;padding:20px">발행 내역 없음</td></tr>'
    return "\n".join(issue_row_html(r) for r in issuances)


def debt_section_html(summary, prev_s, as_of, is_pm):
    R = 'style="text-align:right;white-space:nowrap"'
    def 조(v): return f"{v/1e12:.2f}조원"
    def diff_td(key, light=False):
        if not prev_s: return f'<td {R} style="color:#94a3b8;text-align:right">–</td>'
        d = summary[key] - prev_s.get(key, summary[key])
        if d == 0: return f'<td {R} style="color:#94a3b8;text-align:right">–</td>'
        color = ("#fca5a5" if not light else "#dc2626") if d > 0 else ("#93c5fd" if not light else "#2563eb")
        sign  = "+" if d > 0 else ""
        return f'<td style="color:{color};font-weight:600;text-align:right;white-space:nowrap">{sign}{d/1e12:.2f}조</td>'

    rows = (
        f'<tr style="background:#eef2ff"><td colspan="3" style="font-weight:700;color:#3730a3;font-size:.82rem;padding:7px 12px">📌 사채 (전력채 · 단기사채 · 외화채권)</td></tr>'
        f'<tr><td style="padding-left:18px;color:#475569">전력채</td><td {R}>{조(summary["전력채"])}</td>{diff_td("전력채",True)}</tr>'
        f'<tr><td style="padding-left:18px;color:#475569">단기사채</td><td {R}>{조(summary["단기사채"])}</td>{diff_td("단기사채",True)}</tr>'
        f'<tr><td style="padding-left:18px;color:#475569">외화채권</td><td {R}>{조(summary["외화채권"])}</td>{diff_td("외화채권",True)}</tr>'
        f'<tr style="background:#f1f5f9;font-weight:700"><td>사채 소계</td><td {R}>{조(summary["사채"])}</td>{diff_td("사채",True)}</tr>'
        f'<tr style="background:#f0fdf4"><td colspan="3" style="font-weight:700;color:#166534;font-size:.82rem;padding:7px 12px">📌 사채외 (중장기기업어음 · 은행차입 · 기타)</td></tr>'
        f'<tr><td style="padding-left:18px;color:#475569">중장기기업어음</td><td {R}>{조(summary["중장기기업어음"])}</td>{diff_td("중장기기업어음",True)}</tr>'
        f'<tr><td style="padding-left:18px;color:#475569">은행차입</td><td {R}>{조(summary["은행차입"])}</td>{diff_td("은행차입",True)}</tr>'
        f'<tr style="background:#f1f5f9;font-weight:700"><td>사채외 소계</td><td {R}>{조(summary["사채외"])}</td>{diff_td("사채외",True)}</tr>'
        f'<tr style="background:#0f2a4a;color:#fff"><td style="font-weight:700">총 차입금</td>'
        f'<td style="font-weight:700;font-size:1rem;text-align:right;white-space:nowrap">{조(summary["합계"])}</td>{diff_td("합계")}</tr>'
    )
    note = "" if is_pm else "⏰ 전일 현황 기준 · 당일 발행 반영은 오후 3시 업데이트 예정"
    note_html = f'<p style="font-size:.82rem;color:#64748b;margin-bottom:14px">{note}</p>' if note else ""
    return f"""
<section>
  <h2 style="border-left-color:#6366f1">차입금 현황</h2>
  {note_html}
  <div class="table-wrap" style="max-width:480px">
    <table>
      <thead><tr><th>구분</th><th style="text-align:right">잔액</th><th style="text-align:right">전일비</th></tr></thead>
      <tbody>{rows}</tbody>
    </table>
  </div>
</section>"""


def generate_html(chart, latest, issuances, debt_summary=None, debt_prev=None, debt_as_of=None, debt_is_pm=False, issu_stats=None, kepco_rate_by_date=None, kepco_amt_by_date=None, ktb_rate_by_date=None, ktb_amt_by_date=None):
    today_str = date.today().strftime("%Y년 %m월 %d일")
    latest_dt = latest.get("날짜","–")
    ktb_v  = latest.get("국고채", 0)
    kep_v  = latest.get("한전채", 0)
    sprd_v = round(kep_v - ktb_v, 3) if ktb_v and kep_v else 0

    cards_html = ""
    for label, val, color, unit in [
        ("한전채(3년) 민평", kep_v,  "#f97316", "%"),
        ("국고채(3년) 민평", ktb_v,  "#2563eb", "%"),
        ("스프레드(한전-국고)", sprd_v, "#0d9488", "%p"),
    ]:
        cards_html += f"""
<div class="card" style="border-top:3px solid {color}">
  <div class="cn">{label}</div>
  <div class="cv" style="color:{color}">{val:.3f}<span class="pct">{unit}</span></div>
  <div class="cd">민평평균 · {latest_dt} 기준</div>
</div>"""

    # 전력채 발행금리: 인자로 받은 YTD KOFIA 데이터 우선, 없으면 빈 dict
    _rate     = kepco_rate_by_date or {}
    _amt      = kepco_amt_by_date  or {}
    _ktb_rate = ktb_rate_by_date   or {}
    _ktb_amt  = ktb_amt_by_date    or {}
    labels_json, datasets_json, sprd_min, sprd_max, rate_min, rate_max = build_chart_data(
        chart, _rate, _ktb_rate)
    kepco_amt_json = json.dumps(_amt, ensure_ascii=False)
    ktb_amt_json   = json.dumps(_ktb_amt, ensure_ascii=False)
    today_html  = today_section_html(issuances)
    recent_html = recent_section_html(issuances)
    debt_html   = debt_section_html(debt_summary, debt_prev, debt_as_of, debt_is_pm) if debt_summary else ""
    stats_html  = issu_stats_section_html(issu_stats) if issu_stats else ""

    THEAD = ('<thead><tr>'
             '<th>발행일</th><th>발행기관</th><th>종목명</th>'
             '<th>만기</th><th>발행금액</th><th>발행금리</th>'
             '<th>전일 민평</th><th>발행-민평</th>'
             '</tr></thead>')

    return f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1.0"/>
<title>채권 시장 모니터 — {today_str}</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap" rel="stylesheet"/>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.3/dist/chart.umd.min.js"></script>
<style>
*,*::before,*::after{{box-sizing:border-box;margin:0;padding:0}}
body{{font-family:'Noto Sans KR',sans-serif;background:#f0f4f8;color:#1e293b;min-height:100vh}}
a{{color:#2563eb;text-decoration:none}}a:hover{{text-decoration:underline}}
header{{background:linear-gradient(135deg,#0f2a4a 0%,#1d4ed8 100%);color:#fff;padding:22px 32px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:8px}}
header h1{{font-size:1.55rem;font-weight:700;letter-spacing:-.5px}}
.sub{{font-size:.83rem;opacity:.75}}
main{{max-width:1400px;margin:0 auto;padding:28px 20px}}
section{{margin-bottom:40px}}
h2{{font-size:.98rem;font-weight:700;color:#0f2a4a;border-left:4px solid #2563eb;padding-left:10px;margin-bottom:16px}}
.sub-h3{{font-size:.88rem;font-weight:600;margin-bottom:10px;padding-left:4px}}
.cards{{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:4px}}
@media(max-width:640px){{.cards{{grid-template-columns:1fr}}}}
.card{{background:#fff;border-radius:14px;padding:22px 20px;box-shadow:0 2px 8px rgba(0,0,0,.08)}}
.cn{{font-size:.78rem;color:#64748b;margin-bottom:8px;font-weight:500}}
.cv{{font-size:2rem;font-weight:700}}
.pct{{font-size:1.1rem;font-weight:400;margin-left:2px}}
.cd{{font-size:.72rem;color:#94a3b8;margin-top:6px}}
.today-grid{{display:grid;grid-template-columns:1fr 1fr;gap:20px}}
@media(max-width:900px){{.today-grid{{grid-template-columns:1fr}}}}
.chart-box{{background:#fff;border-radius:14px;padding:24px 20px;box-shadow:0 2px 8px rgba(0,0,0,.08)}}
.note{{margin-top:12px;background:#f8fafc;border:1px solid #e2e8f0;border-radius:8px;padding:10px 14px;font-size:.8rem;color:#475569;line-height:1.7}}
.table-wrap{{background:#fff;border-radius:12px;overflow-x:auto;box-shadow:0 1px 4px rgba(0,0,0,.07)}}
table{{width:100%;border-collapse:collapse;font-size:.85rem}}
thead tr{{background:#0f2a4a;color:#fff}}
thead th{{padding:11px 12px;font-weight:600;text-align:left;white-space:nowrap}}
tbody tr{{border-bottom:1px solid #f1f5f9;transition:background .12s}}
tbody tr:hover{{background:#f8fafc}}
tbody td{{padding:10px 12px;vertical-align:middle}}
td.rate{{font-family:'Courier New',monospace;font-size:.92rem;font-weight:700;color:#1d4ed8}}
.badge{{display:inline-block;padding:2px 9px;border-radius:20px;font-size:.73rem;font-weight:700}}
.badge-3y{{background:#dbeafe;color:#1d4ed8}}
.badge-5y{{background:#dcfce7;color:#15803d}}
.badge-10y{{background:#fce7f3;color:#be185d}}
footer{{text-align:center;padding:22px;font-size:.78rem;color:#94a3b8;line-height:1.9}}
</style>
</head>
<body>
<header>
  <div><h1>📊 채권 시장 모니터</h1>
  <div class="sub">국고채 · 한전채 민평금리 · 공사채 발행현황</div></div>
  <div class="sub">기준일 <strong style="color:#fff">{today_str}</strong></div>
</header>
<main>

<section>
  <h2>직전 영업일 금리 요약</h2>
  <div class="cards">{cards_html}</div>
</section>

{today_html}

{debt_html}

{stats_html}

<section>
  <h2>국고채 3년 · 한전채 3년 · 스프레드 추이 (민평 기준)</h2>
  <div style="display:flex;align-items:center;gap:8px;margin-bottom:14px;flex-wrap:wrap">
    <span style="font-size:.78rem;color:#64748b;font-weight:500">기간</span>
    <button id="btnYtd" onclick="setRange(0)"
      style="padding:5px 16px;border-radius:20px;border:2px solid #2563eb;background:#2563eb;color:#fff;font-size:.8rem;cursor:pointer;font-weight:600;transition:all .15s">연초부터</button>
    <button id="btn4w" onclick="setRange(20)"
      style="padding:5px 16px;border-radius:20px;border:2px solid #cbd5e1;background:#fff;color:#64748b;font-size:.8rem;cursor:pointer;font-weight:600;transition:all .15s">최근 4주</button>
  </div>
  <div class="chart-box">
    <canvas id="yieldChart" style="max-height:440px"></canvas>
  </div>
  <div class="note">
    📌 KOFIA 채권평가 <strong>3사 평균</strong>(나이스피앤아이·한국자산평가·KIS자산평가) 기준 · 영업일만 표시 ·
    <a href="https://www.kofiabond.or.kr" target="_blank" rel="noopener">kofiabond.or.kr</a> 제공
  </div>
</section>

<section>
  <h2>최근 7일 공사채 발행실적 (한전 포함)</h2>
  <div class="table-wrap">
    <table>
      {THEAD}
      <tbody>{recent_html}</tbody>
    </table>
  </div>
  <div class="note">
    📌 전일 민평: 발행일 기준 전 영업일 KOFIA 3사 평균(나이스·한국자산평가·KIS) · 만기는 발행 시점 기준 ·
    <a href="https://www.kofiabond.or.kr" target="_blank">kofiabond.or.kr → 발행시장 → 발행정보종합</a>
  </div>
</section>

</main>
<footer>
  데이터: <a href="https://www.kofiabond.or.kr" target="_blank">금융투자협회 KOFIA</a> ·
  평일 오전 9시·오후 3시 자동 업데이트 · 마지막 생성: {today_str}
</footer>

<script>
const allLabels   = {labels_json};
const allDatasets = {datasets_json};
const RATE_MIN={rate_min}, RATE_MAX={rate_max};
const SPRD_MIN={sprd_min}, SPRD_MAX={sprd_max};
const kepcoAmt = {kepco_amt_json};
const ktbAmt   = {ktb_amt_json};
let activeRange = 0;

function calcAxes(datasets) {{
  const rates   = [...datasets[0].data, ...datasets[1].data].filter(v=>v!=null);
  const spreads = datasets[2].data.filter(v=>v!=null);
  const barRates = [...(datasets[3]?.data||[]), ...(datasets[4]?.data||[])].filter(v=>v!=null);
  const allRates = [...rates, ...barRates];
  const rMin = allRates.length ? Math.min(RATE_MIN, Math.round((Math.min(...allRates)-0.05)*100)/100) : RATE_MIN;
  const rMax = allRates.length ? Math.round((Math.max(...allRates)+0.15)*100)/100 : RATE_MAX;
  const rawSMax = spreads.length ? Math.round((Math.max(...spreads)*1.75)*100)/100 : SPRD_MAX;
  return {{ rMin, rMax, sMin: 0, sMax: Math.max(rawSMax, SPRD_MAX) }};
}}

function buildChartData(n) {{
  const sl = n === 0 ? allLabels.length : Math.min(n, allLabels.length);
  return {{
    labels:   allLabels.slice(-sl),
    datasets: allDatasets.map(ds => ({{...ds, data: ds.data.slice(-sl)}}))
  }};
}}

// 발행금리 라벨을 막대 위에 표시하는 플러그인
const issLabelPlugin = {{
  id: 'issLabel',
  afterDatasetsDraw(chart) {{
    [[3,'rgba(194,65,12,0.9)'],[4,'rgba(67,56,202,0.9)']].forEach(([dsIdx, color]) => {{
      const meta = chart.getDatasetMeta(dsIdx);
      if(!meta || meta.hidden) return;
      const ctx2 = chart.ctx;
      meta.data.forEach((bar, i) => {{
        const v = chart.data.datasets[dsIdx].data[i];
        if(v == null) return;
        ctx2.save();
        ctx2.font = "bold 10px 'Noto Sans KR',sans-serif";
        ctx2.fillStyle = color;
        ctx2.textAlign = 'center';
        ctx2.textBaseline = 'bottom';
        ctx2.fillText(v.toFixed(2)+'%', bar.x, bar.y - 3);
        ctx2.restore();
      }});
    }});
  }}
}};

const initData = buildChartData(0);
const initAxes = calcAxes(initData.datasets);

const yieldChart = new Chart(document.getElementById('yieldChart'), {{
  type:'line',
  data: initData,
  plugins: [issLabelPlugin],
  options:{{
    responsive:true,
    interaction:{{mode:'index',intersect:false}},
    plugins:{{
      legend:{{
        position:'top',
        labels:{{
          font:{{family:"'Noto Sans KR',sans-serif",size:12}},
          usePointStyle:true, padding:20,
          generateLabels(chart) {{
            const items = Chart.defaults.plugins.legend.labels.generateLabels(chart);
            items.forEach(item => {{
              if(item.datasetIndex === 3 || item.datasetIndex === 4) {{
                item.pointStyle = 'rect';
                item.rotation = 0;
              }}
            }});
            return items;
          }}
        }}
      }},
      tooltip:{{
        callbacks:{{
          label: ctx => {{
            if(ctx.datasetIndex===3) {{
              const rate = ctx.parsed.y;
              if(rate == null) return null;
              const lbl = ctx.chart.data.labels[ctx.dataIndex];
              const e = Object.entries(kepcoAmt).find(([k])=>k.slice(5).replace('-','/')=== lbl);
              const amtStr = e ? ` ${{e[1].toLocaleString()}}억원` : '';
              return ` ⚡ 전력채 발행금리: ${{rate.toFixed(3)}}%${{amtStr}}`;
            }}
            if(ctx.datasetIndex===4) {{
              const rate = ctx.parsed.y;
              if(rate == null) return null;
              const lbl = ctx.chart.data.labels[ctx.dataIndex];
              const e = Object.entries(ktbAmt).find(([k])=>k.slice(5).replace('-','/')=== lbl);
              const amtStr = e ? ` ${{e[1].toLocaleString()}}억원` : '';
              return ` 🏛 국고채 3년 발행금리: ${{rate.toFixed(3)}}%${{amtStr}}`;
            }}
            const unit = ctx.datasetIndex===2 ? '%p' : '%';
            return ` ${{ctx.dataset.label}}: ${{ctx.parsed.y.toFixed(3)}}${{unit}}`;
          }}
        }}
      }}
    }},
    scales:{{
      x:{{
        type:'category',
        grid:{{color:'rgba(241,245,249,0.7)'}},
        ticks:{{font:{{family:"'Noto Sans KR'",size:10}},maxRotation:0,autoSkip:true,maxTicksLimit:18}}
      }},
      y:{{
        position:'left',
        min: initAxes.rMin, max: initAxes.rMax,
        title:{{display:true,text:'금리 (%)',font:{{size:10}},color:'#94a3b8'}},
        grid:{{color:'rgba(241,245,249,0.7)'}},
        ticks:{{font:{{family:"'Noto Sans KR'",size:11}},callback:v=>v.toFixed(2)+'%'}}
      }},
      y2:{{
        position:'right',
        min: initAxes.sMin, max: initAxes.sMax,
        title:{{display:true,text:'스프레드 (%p)',font:{{size:10}},color:'#0d9488'}},
        ticks:{{font:{{family:"'Noto Sans KR'",size:10}},color:'#0d9488',callback:v=>v.toFixed(3)+'%p'}},
        grid:{{drawOnChartArea:false}}
      }}
    }}
  }}
}});

function setRange(n) {{
  activeRange = n;
  const d = buildChartData(n);
  const ax = calcAxes(d.datasets);
  yieldChart.data.labels   = d.labels;
  yieldChart.data.datasets = d.datasets;
  yieldChart.options.scales.y.min  = ax.rMin;
  yieldChart.options.scales.y.max  = ax.rMax;
  yieldChart.options.scales.y2.min = ax.sMin;
  yieldChart.options.scales.y2.max = ax.sMax;
  yieldChart.options.scales.x.ticks.maxTicksLimit = n===0 ? 18 : 12;
  yieldChart.update();
  const active = '#2563eb', inactive = '#fff', inactText = '#64748b', inactBrd = '#cbd5e1';
  ['btnYtd','btn4w'].forEach((id,i) => {{
    const on = (i===0 && n===0) || (i===1 && n===20);
    const el = document.getElementById(id);
    el.style.background  = on ? active  : inactive;
    el.style.color       = on ? '#fff'  : inactText;
    el.style.borderColor = on ? active  : inactBrd;
  }});
}}
</script>
</body>
</html>"""


# ── 메인 ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("=" * 52)
    print(f" 채권 시장 모니터 — {date.today()}")
    print("=" * 52)

    chart, latest    = collect_kofia_history()
    issuances        = collect_issuance(days=7)
    d_sum, d_prev, d_as_of, d_pm = collect_debt(issuances)
    issu_stats       = collect_issu_stats()
    kepco_rates, kepco_amts = collect_kepco_rates_ytd()
    ktb_rates,   ktb_amts   = collect_ktb3y_rates_ytd()

    html = generate_html(chart, latest, issuances,
                         debt_summary=d_sum, debt_prev=d_prev,
                         debt_as_of=d_as_of, debt_is_pm=d_pm,
                         issu_stats=issu_stats,
                         kepco_rate_by_date=kepco_rates,
                         kepco_amt_by_date=kepco_amts,
                         ktb_rate_by_date=ktb_rates,
                         ktb_amt_by_date=ktb_amts)
    with open(OUTPUT, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n✅  {OUTPUT} 생성 완료")
    cwd = os.path.dirname(os.path.abspath(__file__))
    print(f"   자동 실행: 0 0,6 * * 1-5 cd {cwd} && python3 update.py >> cron.log 2>&1")
